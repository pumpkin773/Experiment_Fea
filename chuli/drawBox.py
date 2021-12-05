#!/usr/bin/python
# -*- coding:utf-8 -*-
import glob

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

'''每个数据集中，每种方法都取最好的结果'''
def WeMaxdata(eva):
    AF_data, UFS_data = [], []
    CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data = [], [], [], [], [], [], [], []

    for srcFile in glob.glob('../result_data/*.csv'):
        filename = srcFile.split('\\')[-1]
        filename = filename.split('_')
        # print(filename)
        if filename[0] == "Fea":
            pass
        else:
            print(filename)
            data_fram = pd.read_csv(srcFile)
            if filename[0] == "allFea":
                AF_data.append(data_fram[eva].max())
            elif filename[0] == "DBSCAN":
                UFS_data.append(data_fram[eva].max())
            else:
                model_data = data_fram.groupby(['feaSelect'])[eva].max()
                df_data = pd.DataFrame(model_data)
                # print(df_data)
                # print(df_data.loc['Chi2','F1'])
                if np.isnan(df_data.loc['Chi2', eva]):
                    print(df_data.loc['Chi2', eva])
                    CS_data.append(0)
                else:
                    CS_data.append(df_data.loc['Chi2', eva])

                if np.isnan(df_data.loc['Correlate', eva]):
                    print(df_data.loc['Correlate', eva])
                    CFS_data.append(0)
                else:
                    CFS_data.append(df_data.loc['Correlate', eva])

                if np.isnan(df_data.loc['GBDT', eva]):
                    print(df_data.loc['GBDT', eva])
                    TFS_data.append(0)
                else:
                    TFS_data.append(df_data.loc['GBDT', eva])

                if np.isnan(df_data.loc['Muinfor', eva]):
                    print(df_data.loc['Muinfor', eva])
                    IG_data.append(0)
                else:
                    IG_data.append(df_data.loc['Muinfor', eva])

                if np.isnan(df_data.loc['Penalty_l1', eva]):
                    print(df_data.loc['Penalty_l1', eva])
                    L1_data.append(0)
                else:
                    L1_data.append(df_data.loc['Penalty_l1', eva])

                if np.isnan(df_data.loc['Penalty_l1l2', eva]):
                    print(df_data.loc['Penalty_l1l2', eva])
                    L2_data.append(0)
                else:
                    L2_data.append(df_data.loc['Penalty_l1l2', eva])

                if np.isnan(df_data.loc['Variance', eva]):
                    print(df_data.loc['Variance', eva])
                    VT_data.append(0)
                else:
                    VT_data.append(df_data.loc['Variance', eva])

                if np.isnan(df_data.loc['recurElimination', eva]):
                    print(df_data.loc['recurElimination', eva])
                    RFE_data.append(0)
                else:
                    RFE_data.append(df_data.loc['recurElimination', eva])
                # ['AF', 'UFS', 'CS', 'CFS', 'TFS', 'IG', 'L1', 'L2', 'VT', 'RFE']
                # data = [AF_data, UFS_data, CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data]
            data = [AF_data, UFS_data, CS_data, CFS_data, TFS_data, IG_data, L2_data, VT_data, RFE_data]

            '''model_index = data_fram['F1'].idxmin()
                    print(data_fram.iat[model_index, 9])
                    allFea_data.append(data_fram['F1'].min())
                    '''
    # data = [AF_data, UFS_data, CS_data, CFS_data, TFS_data, IG_data, L2_data, VT_data, RFE_data]
    data = [UFS_data, AF_data, VT_data, CS_data, CFS_data, IG_data, RFE_data, TFS_data, L2_data]
    return data

def drawtimeBOX(ml,data,color_list):

    labels = 'Our approach', 'VT', 'CS', 'CFS', 'IG', 'RFE', 'TFS', 'RFS'  # 图例
    bp = plt.boxplot(data, notch=False, labels=labels, patch_artist=False,
                boxprops={'color': 'black', 'linewidth': '1.0'},
                capprops={'color': 'black', 'linewidth': '1.0'},
                medianprops={'color':'black','linewidth':'1.5'})
    #边框 color=c, linewidth=2
    #facecolor整个箱
    for box, c in zip(bp['boxes'], color_list):
        # 箱体边框颜色
        box.set(color=c, linewidth=1)
    #[bp['boxes'][i].set(color=color_list[i], linewidth=2, alpha=0.7) for i in range(10)]
    # 当然也可以按照for循环进行挨个设置，也可以列表表达式
    if ml != 'NO':
        plt.title(ml, fontsize=13)
    plt.ylabel('Cost', fontsize=10)
    # plt.xticks(fontsize=10)
    # plt.yticks(fontsize=10)
    plt.show()
    return np.array(data)

def timedata(eva):
    AF_data, UFS_data = [], []
    CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data = [], [], [], [], [], [], [], []

    for srcFile in glob.glob('../result_data/*.csv'):
        filename = srcFile.split('\\')[-1]
        filename = filename.split('_')
        # print(filename)
        if filename[0] == "Fea":
            pass
        else:
            print(filename)
            data_fram = pd.read_csv(srcFile)
            if filename[0] == "DBSCAN":
                model_index = data_fram[eva].idxmax()
                print(data_fram.iat[model_index, 14])
                UFS_data.append(data_fram.iat[model_index, 14])

            elif filename[0] == "general":
                model_index = data_fram.groupby(['feaSelect'])[eva].idxmax()
                model_data = data_fram.iloc[model_index, :]
                df_data = pd.DataFrame(model_data)
                CS_data.append(df_data[df_data['feaSelect'].isin(['Chi2'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Chi2'])].iat[0, 14])
                CFS_data.append(df_data[df_data['feaSelect'].isin(['Correlate'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Correlate'])].iat[0, 14])
                TFS_data.append(df_data[df_data['feaSelect'].isin(['GBDT'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['GBDT'])].iat[0, 14])
                IG_data.append(df_data[df_data['feaSelect'].isin(['Muinfor'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Muinfor'])].iat[0, 14])
                L1_data.append(df_data[df_data['feaSelect'].isin(['Penalty_l1'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Penalty_l1'])].iat[0, 14])
                L2_data.append(df_data[df_data['feaSelect'].isin(['Penalty_l1l2'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Penalty_l1l2'])].iat[0, 14])
                VT_data.append(df_data[df_data['feaSelect'].isin(['Variance'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['Variance'])].iat[0, 14])
                RFE_data.append(df_data[df_data['feaSelect'].isin(['recurElimination'])].iat[0, 14])
                print(df_data[df_data['feaSelect'].isin(['recurElimination'])].iat[0, 14])
    data = [UFS_data, VT_data, CS_data, CFS_data, IG_data, RFE_data, TFS_data, L2_data]
    return data

def AUC_F1MCC(eva_index):
    AF_data, UFS_data = [], []
    CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data = [], [], [], [], [], [], [], []

    for srcFile in glob.glob('../result_data/*.csv'):
        filename = srcFile.split('\\')[-1]
        filename = filename.split('_')
        # print(filename)
        if filename[0] == "Fea":
            pass
        else:
            print(filename)
            data_fram = pd.read_csv(srcFile)
            if filename[0] == "allFea":
                model_index = data_fram['AUC'].idxmax()
                AF_data.append(data_fram.iat[model_index, eva_index-1])
            elif filename[0] == "DBSCAN":
                model_index = data_fram['AUC'].idxmax()
                UFS_data.append(data_fram.iat[model_index, eva_index])
            elif filename[0] == "general":
                model_index = data_fram.groupby(['feaSelect'])['AUC'].idxmax()
                model_data = data_fram.iloc[model_index, :]
                df_data = pd.DataFrame(model_data)
                CS_data.append(df_data[df_data['feaSelect'].isin(['Chi2'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Chi2'])].iat[0, eva_index])
                CFS_data.append(df_data[df_data['feaSelect'].isin(['Correlate'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Correlate'])].iat[0, eva_index])
                TFS_data.append(df_data[df_data['feaSelect'].isin(['GBDT'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['GBDT'])].iat[0, eva_index])
                IG_data.append(df_data[df_data['feaSelect'].isin(['Muinfor'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Muinfor'])].iat[0, eva_index])
                L1_data.append(df_data[df_data['feaSelect'].isin(['Penalty_l1'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Penalty_l1'])].iat[0, eva_index])
                L2_data.append(df_data[df_data['feaSelect'].isin(['Penalty_l1l2'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Penalty_l1l2'])].iat[0, eva_index])
                VT_data.append(df_data[df_data['feaSelect'].isin(['Variance'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['Variance'])].iat[0, eva_index])
                RFE_data.append(df_data[df_data['feaSelect'].isin(['recurElimination'])].iat[0, eva_index])
                print(df_data[df_data['feaSelect'].isin(['recurElimination'])].iat[0, eva_index])
    data = [UFS_data, AF_data, VT_data, CS_data, CFS_data, IG_data, RFE_data, TFS_data, L2_data]
    return data


def drawBOX(ml,data,color_list,ylabel):
    #matrix = np.array([allFea_data, DBSCAN_data, CS_data, CFS_data, TFS_data, IG_data, L1_data, L1L2_data, VT_data, RFE_data])
    #data = [allFea_data, DBSCAN_data, CS_data, CFS_data, TFS_data, IG_data, L1_data, L1L2_data, VT_data, RFE_data]
    # print(len(allFea_data),len(DBSCAN_data),len(CS_data))
    # print(len(CFS_data), len(TFS_data), len(IG_data))
    # print(len(L1_data),len(L1L2_data),len(VT_data),len(RFE_data))
    labels = 'Our approach', 'AF', 'VT', 'CS', 'CFS', 'IG', 'RFE', 'TFS', 'RFS' # 图例
    bp = plt.boxplot(data, notch=False, labels=labels, patch_artist=False,
                boxprops={'color': 'black', 'linewidth': '1.0'},
                capprops={'color': 'black', 'linewidth': '1.0'},
                medianprops={'color':'black','linewidth':'1.5'})
    #边框 color=c, linewidth=2
    #facecolor整个箱
    for box, c in zip(bp['boxes'], color_list):
        # 箱体边框颜色
        box.set(color=c, linewidth=1)
    #[bp['boxes'][i].set(color=color_list[i], linewidth=2, alpha=0.7) for i in range(10)]
    # 当然也可以按照for循环进行挨个设置，也可以列表表达式
    if ml != 'NO':
        plt.title(ml, fontsize=13)
    plt.ylabel(ylabel, fontsize=10)
    # plt.xticks(fontsize=10)
    # plt.yticks(fontsize=10)
    plt.show()
    return np.array(data)

'构造降序排序矩阵'
def rank_matrix(matrix):
    cnum = matrix.shape[1]
    rnum = matrix.shape[0]
    ## 升序排序索引
    sorts = np.argsort(matrix)
    for i in range(rnum):
        k = 1
        n = 0
        flag = False
        nsum = 0
        for j in range(cnum):
            n = n + 1
            ## 相同排名评分序值
            if j < 2 and matrix[i, sorts[i, j]] == matrix[i, sorts[i, j + 1]]:
                flag = True;
                k = k + 1;
                nsum += j + 1;
            elif (j == 2 or (j < 2 and matrix[i, sorts[i, j]] != matrix[i, sorts[i, j + 1]])) and flag:
                nsum += j + 1
                flag = False;
                for q in range(k):
                    matrix[i, sorts[i, j - k + q + 1]] = nsum / k
                k = 1
                flag = False
                nsum = 0
            else:
                matrix[i, sorts[i, j]] = j + 1
                continue
    return matrix

"""
    Friedman检验
    参数：数据集个数n, 算法种数k, 排序矩阵rank_matrix(k x n)
    函数返回检验结果（对应于排序矩阵列顺序的一维数组）
"""
def friedman(n, k, rank_matrix):
    # 计算每一列的排序和
    sumr = sum(list(map(lambda x: np.mean(x) ** 2, rank_matrix.T)))
    result = 12 * n / (k * (k + 1)) * (sumr - k * (k + 1) ** 2 / 4)
    result = (n - 1) * result / (n * (k - 1) - result)
    return result

"""
    Nemenyi检验
    参数：数据集个数n, 算法种数k, 排序矩阵rank_matrix(k x n)
    函数返回CD值
"""
def nemenyi(n, k, q):
    return q * (np.sqrt(k * (k + 1) / (6 * n)))

def main():
    ML = ['NB', 'LR', 'SVM', 'RF', 'DT', 'KNN']
    # 'select_Fea_time'
    EVA = ['AUC']
    # F1---23 mcc----27
    # EVA = ['F1','MCC']
    ML = ['NO']
    for ml in ML:
        for eva in EVA:
            # data = WeMaxdata(eva)
            data = timedata(eva)
            if eva == 'F1':
                eva_index = 23
            elif eva == 'MCC':
                eva_index = 27
            # data = AUC_F1MCC(eva_index)
            color_list = ['black' for _ in range(8)]
            drawtimeBOX(ml, data, color_list)
            # drawBOX(ml, data, color_list, eva)
            matrix = np.array(data)
            matrix_r = rank_matrix(matrix.T)
            Friedman = friedman(9, 8, matrix_r)
            print("================")
            print(Friedman)
            CD = nemenyi(9, 8, 3.031)#9, 10, 3.164 #9, 9, 3.102
            ##画CD图
            rank_x = list(map(lambda x: np.mean(x), matrix))
            print(rank_x)
            #['AF', 'UFS', 'CS', 'CFS', 'TFS', 'IG', 'L1', 'RFS', 'VT', 'RFE']
            # name_y = ['Our approach', 'AF', 'VT', 'CS', 'CFS', 'IG', 'RFE', 'TFS', 'RFS']
            name_y = ['Our approach', 'VT', 'CS', 'CFS', 'IG', 'RFE', 'TFS', 'RFS']
            min_ = [x for x in rank_x - CD / 2]
            max_ = [x for x in rank_x + CD / 2]
            print(min_)
            print(max_)
            # plt.title("Friedman")
            plt.scatter(rank_x, name_y)
            plt.hlines(name_y, min_, max_)
            plt.axvline(x=min_[0], ls="--", c="green")
            plt.axvline(x=max_[0], ls="--", c="green")
            plt.show()
            lableM = []
            for i in range(len(min_)):
                if min_[i] < min_[0] and max_[i] < min_[0]:
                    lableM.append('red')
                elif min_[i] > max_[0] and max_[i] > max_[0]:
                    lableM.append('green')
                else:
                    lableM.append('blue')
            drawtimeBOX(ml, data, lableM)
            # drawBOX(ml, data, lableM, eva)


def finall_box():
    wb = openpyxl.load_workbook("../data/final result.xlsx")
    # 获取workbook中所有的表格
    sheets = wb.sheetnames
    print(sheets)
    # 循环遍历所有sheet
    AF_data, Ourapproach_data = [], []
    CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data = [], [], [], [], [], [], [], []
    # for i in range(6):
    # 25----AUC；15---特征选择时间；27--MCC
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
        for r in range(1, sheet.max_row + 1):
            if r == 1:
                print(str(sheet.cell(row=r, column=28).value))
                pass
                # print('\n' + ''.join([str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
            else:
                if str(sheet.cell(row=r, column=5).value) == "Variance":
                    VT_data.append(float(sheet.cell(row=r, column=28).value))
                    # print(float(sheet.cell(row=r, column=25).value))
                elif str(sheet.cell(row=r, column=5).value) == "All":
                    AF_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "Penalty_l1l2":
                    L2_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "recurElimination":
                    RFE_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "Muinfor":
                    IG_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "Chi2":
                    CS_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "Correlate":
                    CFS_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "DBSCAN":
                    Ourapproach_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "Penalty_l1":
                    L1_data.append(float(sheet.cell(row=r, column=28).value))
                elif str(sheet.cell(row=r, column=5).value) == "GBDT":
                    TFS_data.append(float(sheet.cell(row=r, column=28).value))
    # AF_data,
    data = [AF_data, Ourapproach_data, CS_data, CFS_data, TFS_data,
            IG_data, L1_data, L2_data, VT_data, RFE_data]
    ml = 'NO'
    color_list = ['black' for _ in range(10)]
    print(data)
    drawBOX(ml, data, color_list)
    matrix = np.array(data)
    matrix_r = rank_matrix(matrix.T)
    # Friedman = friedman(9, 10, matrix_r)
    CD = nemenyi(9, 10, 3.164)
    # CD = nemenyi(9, 9, 3.102)
    ##画CD图
    rank_x = list(map(lambda x: np.mean(x), matrix))
    print(rank_x)
    # 我们的方法UFS
    name_y = ['AF', 'UFS', 'CS', 'CFS', 'TFS', 'IG', 'L1', 'RFS', 'VT', 'RFE']
    # name_y = [ 'UFS', 'CS', 'CFS', 'TFS', 'IG', 'L1', 'L2', 'VT', 'RFE']
    min_ = [x for x in rank_x - CD / 2]
    max_ = [x for x in rank_x + CD / 2]
    print(min_)
    print(max_)
    plt.title("Friedman")
    plt.scatter(rank_x, name_y)
    plt.hlines(name_y, min_, max_)
    plt.axvline(x=min_[1], ls="--", c="green")
    plt.axvline(x=max_[1], ls="--", c="green")
    # plt.axvline(x=min_[0], ls="--", c="green")
    # plt.axvline(x=max_[0], ls="--", c="green")
    plt.show()
    lableM = []
    for i in range(len(min_)):
        # 这里从RFE分组
        if min_[i] < min_[1] and max_[i] < min_[1]:
            # if min_[i] < min_[0] and max_[i] < min_[0]:
            lableM.append('red')
        elif min_[i] > max_[1] and max_[i] > max_[1]:
            # elif min_[i] > max_[0] and max_[i] > max_[0]:
            lableM.append('green')
        else:
            lableM.append('blue')
    drawBOX(ml, data, lableM)

def jaccard_sim(a, b):
    unions = len(set(a).union(set(b)))
    intersections = len(set(a).intersection(set(b)))
    return 1. * intersections / unions

def cal_jaccard():
    wb = openpyxl.load_workbook("../data/final Fea result.xlsx")
    sheets = wb.sheetnames
    print(sheets)
    AF_data, UFS_data = [], []
    CS_data, CFS_data, TFS_data, IG_data, L1_data, L2_data, VT_data, RFE_data = [], [], [], [], [], [], [], []
    #'UFS', 'CS', 'CFS', 'TFS', 'IG', 'L1', 'L2', 'VT', 'RFE'
    #UFS_AF, UFS_CS, UFS_CFS, UFS_TFS, UFS_IG, UFS_L1, UFS_L2, UFS_VT, UFS_RFE = [], [], [], [], [], [], [], [], []
    jaccard_matrix=[]
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        print('\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
        for r in range(1, sheet.max_row + 1):
            if r == 1:
                print(str(sheet.cell(row=r, column=8).value))
            else:
                if str(sheet.cell(row=r, column=1).value) == "Variance":
                    VT_data = sheet.cell(row=r, column=8).value
                    VT_data = VT_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "All":
                    AF_data = sheet.cell(row=r, column=8).value
                    AF_data = AF_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "Penalty_l1l2":
                    L2_data = sheet.cell(row=r, column=8).value
                    L2_data = L2_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "recurElimination":
                    RFE_data = sheet.cell(row=r, column=8).value
                    RFE_data = RFE_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "Muinfor":
                    IG_data = sheet.cell(row=r, column=8).value
                    IG_data = IG_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "Chi2":
                    CS_data = sheet.cell(row=r, column=8).value
                    CS_data = CS_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "Correlate":
                    CFS_data = sheet.cell(row=r, column=8).value
                    CFS_data = CFS_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "DBSCAN":
                    UFS_data = sheet.cell(row=r, column=8).value
                    UFS_data = UFS_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "Penalty_l1":
                    L1_data = sheet.cell(row=r, column=8).value
                    L1_data = L1_data.split(',')[:-1]
                elif str(sheet.cell(row=r, column=1).value) == "GBDT":
                    TFS_data = sheet.cell(row=r, column=8).value
                    TFS_data = TFS_data.split(',')[:-1]
        data = [AF_data,CS_data, CFS_data, TFS_data,IG_data, L1_data, L2_data, VT_data, RFE_data]
        #UFS_AF, UFS_CS, UFS_CFS, UFS_TFS, UFS_IG, UFS_L1, UFS_L2, UFS_VT, UFS_RFE
        jaccard_project_matrix = []
        for i in range(len(data)):
            ja = jaccard_sim(UFS_data, data[i])
            # print(ja)
            jaccard_project_matrix.append(ja)
        # print(jaccard_project_matrix)
        jaccard_matrix.append(jaccard_project_matrix)
        # print(jaccard_matrix)
    jaccard_dataframe = pd.DataFrame(jaccard_matrix, index=['ant', 'commons', 'derby',
                                                            'jmeter','mvn','lucence',
                                                            'cass','phoenix','tomcat'],
                                     columns=['UFS_AF', 'UFS_CS', 'UFS_CFS','UFS_TFS',
                                              'UFS_IG','UFS_L1','UFS_RFS','UFS_VT','UFS_RFE'])
    print(jaccard_dataframe)
    jaccard_dataframe.to_csv('../data/jaccard_result.csv')

def Feamap(fea):
    print(fea)
    fileMap = "../data/feature id mapping.csv"
    origin_data_frame2 = pd.read_csv(fileMap)
    pro_list = list(origin_data_frame2['id in the program'])
    par_list = list(origin_data_frame2['name in paper'])
    print(pro_list)
    print(par_list)
    c = fea.split('F')[1]
    paperFea = 0
    for ip in range(len(pro_list)):
        if np.isnan(pro_list[ip]):
            continue
        if int(c.split('-')[0]) == int(pro_list[ip]):
            paperFea =par_list[ip]
            break
    if paperFea:
        print(paperFea)
        return paperFea
    else:
        return 'NO'


def cal_frequency():
    wb = openpyxl.load_workbook("../data/final Fea result.xlsx")
    sheets = wb.sheetnames
    print(sheets)
    project_fea_dic = {}
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        print('\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
        fea_dic={}
        for r in range(1, sheet.max_row + 1):
            if r == 1:
                print(str(sheet.cell(row=r, column=8).value))
            else:
                fea_data = sheet.cell(row=r, column=8).value
                fea_data = fea_data.split(',')[:-1]
                for fea in fea_data:
                    pFea = Feamap(fea)
                    # pFea = fea
                    if pFea == 'NO':
                        continue
                    if not fea_dic.__contains__(pFea):
                        fea_dic[pFea] = 1
                    fea_dic[pFea] += 1
        project_fea_dic[sheets[i]] = fea_dic
        print(fea_dic)
    project_fea_dataframe = pd.DataFrame(project_fea_dic)
    project_fea_dataframe = pd.DataFrame(project_fea_dataframe.values.T, index=project_fea_dataframe.columns, columns=project_fea_dataframe.index)
    print(project_fea_dataframe)
    project_fea_dataframe.to_csv('../data/fea_frequency.csv')


'''每个数据集中，我们自己的方法不同策略都取最好的结果对比'''
def We3Maxdata(eva_index):
    std_data, inclass_corr_data, LDF_data = [], [], []

    for srcFile in glob.glob('../result_data/*.csv'):
        filename = srcFile.split('\\')[-1]
        filename = filename.split('_')
        # print(filename)
        if filename[0] == "Fea":
            pass
        else:
            print(filename)
            data_fram = pd.read_csv(srcFile)
            if filename[0] == "DBSCAN":
                model_index = data_fram.groupby(['select_method'])['AUC'].idxmax()
                model_data = data_fram.iloc[model_index, :]
                df_data = pd.DataFrame(model_data)
                # AUC--24 , F1---23 mcc----27
                std_data.append(df_data[df_data['select_method'].isin(['std'])].iat[0, eva_index])
                print(df_data[df_data['select_method'].isin(['std'])].iat[0, eva_index])
                inclass_corr_data.append(df_data[df_data['select_method'].isin(['inclass_corr'])].iat[0, eva_index])
                print(df_data[df_data['select_method'].isin(['inclass_corr'])].iat[0, eva_index])
                LDF_data.append(df_data[df_data['select_method'].isin(['LDF'])].iat[0, eva_index])
                print(df_data[df_data['select_method'].isin(['LDF'])].iat[0, eva_index])
            else:
                pass
    # data = [AF_data, UFS_data, CS_data, CFS_data, TFS_data, IG_data, L2_data, VT_data, RFE_data]
    data = [LDF_data, inclass_corr_data, std_data]
    return data

def draw3BOX(data,color_list,ylabel):
    labels = 'LDR', 'FCR', 'FDR' # 图例
    bp = plt.boxplot(data, notch=False, labels=labels, patch_artist=False,
                boxprops={'color': 'black', 'linewidth': '1.0'},
                capprops={'color': 'black', 'linewidth': '1.0'},
                medianprops={'color':'black','linewidth':'1.5'})
    #边框 color=c, linewidth=2
    #facecolor整个箱
    for box, c in zip(bp['boxes'], color_list):
        # 箱体边框颜色
        box.set(color=c, linewidth=1)
    #[bp['boxes'][i].set(color=color_list[i], linewidth=2, alpha=0.7) for i in range(10)]
    # 当然也可以按照for循环进行挨个设置，也可以列表表达式
    plt.ylabel(ylabel, fontsize=10)
    # plt.xticks(fontsize=10)
    # plt.yticks(fontsize=10)
    plt.show()
    return np.array(data)

def main3():
    EVA = ['AUC', 'F1', 'MCC', 'Cost (:s)']
    # AUC--24 , F1---23 mcc----27, cost--14
    for eva in EVA:
        if eva == 'F1':
            eva_index = 23
        elif eva == 'MCC':
            eva_index = 27
        elif eva == 'AUC':
            eva_index = 24
        elif eva == 'Cost (:s)':
            eva_index = 14
        data = We3Maxdata(eva_index)
        color_list = ['black' for _ in range(3)]
        draw3BOX(data, color_list, eva)
        matrix = np.array(data)
        matrix_r = rank_matrix(matrix.T)
        Friedman = friedman(9, 3, matrix_r)
        print("================")
        print(Friedman)
        CD = nemenyi(9, 3, 2.344)
        ##画CD图
        rank_x = list(map(lambda x: np.mean(x), matrix))
        print(rank_x)
        #std_data, inclass_corr_data, LDF_data
        #name_y = ['FDR', 'FCR', 'LDR']
        name_y = ['LDR', 'FCR', 'FDR']
        min_ = [x for x in rank_x - CD / 2]
        max_ = [x for x in rank_x + CD / 2]
        print(min_)
        print(max_)
        plt.scatter(rank_x, name_y)
        plt.hlines(name_y, min_, max_)
        plt.show()
        lableM = []
        for i in range(len(min_)):
            if min_[i] < min_[0] and max_[i] < min_[0]:
                lableM.append('red')
            elif min_[i] > max_[0] and max_[i] > max_[0]:
                lableM.append('green')
            else:
                lableM.append('blue')
        draw3BOX(data, lableM, eva)

'''每个数据集中，我们自己的方法不同策略下的不同机器学习模型结果'''
def MLdata(eva_index):
    #NB、LR、KNN、DT、RF、SVM
    NB_data, LR_data, KNN_data, DT_data, RF_data, SVM_data = [], [], [], [], [], []

    for srcFile in glob.glob('../result_data/*.csv'):
        filename = srcFile.split('\\')[-1]
        filename = filename.split('_')
        # print(filename)
        if filename[0] == "Fea":
            pass
        else:
            print(filename)
            data_fram = pd.read_csv(srcFile)
            if filename[0] == "DBSCAN":
                # std_data, inclass_corr_data, LDF_data
                # ['FDR', 'FCR', 'LDR']
                if filename[2] == "ant":
                    model_data = data_fram[data_fram['select_method'].isin(['inclass_corr'])]
                elif filename[2] == "cass":
                    model_data = data_fram[data_fram['select_method'].isin(['LDF'])]
                elif filename[2] == "commons":
                    model_data = data_fram[data_fram['select_method'].isin(['std'])]
                elif filename[2] == "derby":
                    model_data = data_fram[data_fram['select_method'].isin(['std'])]
                elif filename[2] == "jmeter":
                    model_data = data_fram[data_fram['select_method'].isin(['inclass_corr'])]
                elif filename[2] == "lucence":
                    model_data = data_fram[data_fram['select_method'].isin(['LDF'])]
                elif filename[2] == "mvn":
                    model_data = data_fram[data_fram['select_method'].isin(['LDF'])]
                elif filename[2] == "phoenix":
                    model_data = data_fram[data_fram['select_method'].isin(['LDF'])]
                elif filename[2] == "tomcat":
                    model_data = data_fram[data_fram['select_method'].isin(['LDF'])]
                df_data = pd.DataFrame(model_data)
                #NB_data, LR_data, KNN_data, DT_data, RF_data, SVM_data = [], [], [], [], [], []
                NB_data.append(df_data[df_data['MLmodel'].isin(['NB'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['NB'])].iat[0, eva_index])
                LR_data.append(df_data[df_data['MLmodel'].isin(['LR'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['LR'])].iat[0, eva_index])
                KNN_data.append(df_data[df_data['MLmodel'].isin(['KNN'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['KNN'])].iat[0, eva_index])
                DT_data.append(df_data[df_data['MLmodel'].isin(['DT'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['DT'])].iat[0, eva_index])
                RF_data.append(df_data[df_data['MLmodel'].isin(['RF'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['RF'])].iat[0, eva_index])
                SVM_data.append(df_data[df_data['MLmodel'].isin(['SVM'])].iat[0, eva_index])
                print(df_data[df_data['MLmodel'].isin(['SVM'])].iat[0, eva_index])
            else:
                pass
    data = [NB_data, LR_data, KNN_data, DT_data, RF_data, SVM_data]
    return data

def drawMLBOX(data,color_list,ylabel):
    labels = 'NB', 'LR', 'KNN', 'DT', 'RF', 'SVM' # 图例
    bp = plt.boxplot(data, notch=False, labels=labels, patch_artist=False,
                boxprops={'color': 'black', 'linewidth': '1.0'},
                capprops={'color': 'black', 'linewidth': '1.0'},
                medianprops={'color':'black','linewidth':'1.5'})
    #边框 color=c, linewidth=2
    #facecolor整个箱
    for box, c in zip(bp['boxes'], color_list):
        # 箱体边框颜色
        box.set(color=c, linewidth=1)
    #[bp['boxes'][i].set(color=color_list[i], linewidth=2, alpha=0.7) for i in range(10)]
    # 当然也可以按照for循环进行挨个设置，也可以列表表达式
    plt.ylabel(ylabel, fontsize=10)
    # plt.xticks(fontsize=10)
    # plt.yticks(fontsize=10)
    plt.show()
    return np.array(data)

def mainML():
    EVA = ['AUC', 'F1', 'MCC']
    # AUC--24 , F1---23 mcc----27
    for eva in EVA:
        if eva == 'F1':
            eva_index = 23
        elif eva == 'MCC':
            eva_index = 27
        elif eva == 'AUC':
            eva_index = 24
        data = MLdata(eva_index)
        color_list = ['black' for _ in range(6)]
        drawMLBOX(data, color_list, eva)
        matrix = np.array(data)
        matrix_r = rank_matrix(matrix.T)
        Friedman = friedman(9, 6, matrix_r)
        print("================")
        print(Friedman)
        CD = nemenyi(9, 6, 2.850)
        ##画CD图
        rank_x = list(map(lambda x: np.mean(x), matrix))
        print(rank_x)
        #std_data, inclass_corr_data, LDF_data
        name_y = ['NB', 'LR', 'KNN', 'DT', 'RF', 'SVM']
        min_ = [x for x in rank_x - CD / 2]
        max_ = [x for x in rank_x + CD / 2]
        print(min_)
        print(max_)
        plt.scatter(rank_x, name_y)
        plt.hlines(name_y, min_, max_)
        plt.show()
        lableM = []
        for i in range(len(min_)):
            if min_[i] < min_[0] and max_[i] < min_[0]:
                lableM.append('red')
            elif min_[i] > max_[0] and max_[i] > max_[0]:
                lableM.append('green')
            else:
                lableM.append('blue')
        drawMLBOX(data, lableM, eva)

if __name__ == "__main__":
    #cal_jaccard()
    # cal_frequency()
    # main()
    # main3()
    mainML()


